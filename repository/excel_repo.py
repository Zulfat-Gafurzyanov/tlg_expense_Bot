from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import DATA_DIR

# Путь к единственному файлу со всеми данными
EXPENSES_FILE: Path = DATA_DIR / "expenses.xlsx"


# --- Вспомогательные функции ---


def _get_current_sheet_name() -> str:
    """Возвращает имя листа для текущего месяца, например '2026_02'."""
    now = datetime.now()
    return f"{now.year}_{now.month:02d}"


def _ensure_file_and_sheet() -> tuple[Workbook, Worksheet]:
    """
    Гарантирует существование файла и листа текущего месяца.
    Если файл не существует — создаёт его.
    Если лист для текущего месяца не существует — добавляет его с заголовком.
    Возвращает workbook и лист текущего месяца.
    """
    sheet_name = _get_current_sheet_name()

    if not EXPENSES_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["ID", "Дата", "Категория", "Сумма", "Комментарий"])
        wb.save(EXPENSES_FILE)
        return wb, ws

    wb = load_workbook(EXPENSES_FILE)

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["ID", "Дата", "Категория", "Сумма", "Комментарий"])
        wb.save(EXPENSES_FILE)
        return wb, ws

    ws = wb[sheet_name]
    return wb, ws


def _get_next_id(ws: Worksheet) -> int:
    """
    Определяет следующий ID как max(существующих ID) + 1.
    Это гарантирует что ID всегда растёт вверх,
    даже если промежуточные записи были удалены.
    """
    if ws.max_row <= 1:
        return 1

    max_id = 0
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=1).value
        if cell_value is not None and cell_value > max_id:
            max_id = cell_value

    return max_id + 1


# --- Публичный API репозитория ---


def add_expense(category: str, amount: float, comment: str | None = None) -> int:
    """
    Добавляет новую запись расхода в лист текущего месяца.
    Возвращает ID созданной записи.
    """
    wb, ws = _ensure_file_and_sheet()

    new_id = _get_next_id(ws)
    date_str = datetime.now().strftime("%d.%m.%Y")

    ws.append([new_id, date_str, category, amount, comment or ""])
    wb.save(EXPENSES_FILE)

    return new_id


def get_all_expenses() -> list[dict]:
    """
    Возвращает список всех записей из листа текущего месяца.
    Каждая запись — словарь: {id, date, category, amount, comment}.
    """
    if not EXPENSES_FILE.exists():
        return []

    sheet_name = _get_current_sheet_name()
    wb = load_workbook(EXPENSES_FILE)

    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]

    expenses = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        expenses.append({
            "id": row[0],
            "date": row[1],
            "category": row[2],
            "amount": row[3],
            "comment": row[4] if len(row) > 4 and row[4] else "",
        })

    return expenses


def delete_expense(expense_id: int) -> bool:
    """
    Удаляет строку с указанным ID физически.
    ID остальных записей НЕ меняются.
    Возвращает True если запись найдена и удалена, иначе False.
    """
    if not EXPENSES_FILE.exists():
        return False

    wb, ws = _ensure_file_and_sheet()

    # Ищем строку с нужным ID
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == expense_id:
            target_row = row_idx
            break

    if target_row is None:
        return False

    # Удаляем строку, ID не пересчитываем
    ws.delete_rows(target_row)
    wb.save(EXPENSES_FILE)
    return True


def update_expense_category(expense_id: int, new_category: str) -> bool:
    """
    Обновляет категорию записи по ID в листе текущего месяца.
    Возвращает True если запись найдена и обновлена, иначе False.
    """
    if not EXPENSES_FILE.exists():
        return False

    wb, ws = _ensure_file_and_sheet()

    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == expense_id:
            ws.cell(row=row_idx, column=3).value = new_category
            wb.save(EXPENSES_FILE)
            return True

    return False


def update_expense_amount(expense_id: int, new_amount: float) -> bool:
    """
    Обновляет сумму записи по ID в листе текущего месяца.
    Возвращает True если запись найдена и обновлена, иначе False.
    """
    if not EXPENSES_FILE.exists():
        return False

    wb, ws = _ensure_file_and_sheet()

    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == expense_id:
            ws.cell(row=row_idx, column=4).value = new_amount
            wb.save(EXPENSES_FILE)
            return True

    return False


def update_expense_comment(expense_id: int, new_comment: str) -> bool:
    """
    Обновляет комментарий записи по ID в листе текущего месяца.
    Возвращает True если запись найдена и обновлена, иначе False.
    """
    if not EXPENSES_FILE.exists():
        return False

    wb, ws = _ensure_file_and_sheet()

    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == expense_id:
            ws.cell(row=row_idx, column=5).value = new_comment
            wb.save(EXPENSES_FILE)
            return True

    return False


def get_current_month_label() -> str:
    """Возвращает читаемую метку текущего месяца, например 'Февраль 2026'."""
    months = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
    ]
    now = datetime.now()
    return f"{months[now.month - 1]} {now.year}"


def get_file_path() -> Path:
    """Возвращает путь к файлу expenses.xlsx (для экспорта)."""
    return EXPENSES_FILE